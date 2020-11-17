import openpyxl
from bs4 import BeautifulSoup
from urllib.request import urlopen
import re


workbook = openpyxl.load_workbook('AllDataSets.xlsx')
worksheet = workbook['Import Alerts']

firm_name = str(input('Please input a firm name: '))
resultset = []
for iterator in range(2,worksheet.max_row+1):
   try:
       values = worksheet.cell(row=iterator,column=4).value
       link = worksheet.cell(row=iterator,column=4).hyperlink

       client = urlopen(link.display)
       page = client.read()
       client.close()

       page_container = BeautifulSoup(page, 'html.parser')
       container = page_container.findAll('div', {'class': 'div-info'})

       for firms in container:
           content = firms.find_all('div')
           try:
               if re.search(firm_name, content[0].text, re.IGNORECASE):
                   print("Firm: ", firm_name, "\n\t found in import alert: ", values, "\n",content[1].text)
                   resultset.append([firm_name,values,content[1].text])

           except Exception:
               pass
   except:
       pass

for thingy in resultset:
    print(thingy)