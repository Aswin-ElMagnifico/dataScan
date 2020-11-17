import openpyxl

workbook = openpyxl.load_workbook('AllDataSets.xlsx')

worksheet = workbook['Import Alerts']

firm_name = str(input('Please input a firm name: '))

for iterator in range(2,worksheet.max_row+1):
   try:
       values = worksheet.cell(row=iterator,column=4).value
       link = worksheet.cell(row=iterator,column=4).hyperlink

   except:
       pass
