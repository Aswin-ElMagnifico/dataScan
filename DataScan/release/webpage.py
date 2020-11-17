from flask import Flask, render_template,request
import openpyxl
from bs4 import BeautifulSoup
from urllib.request import urlopen
import re

workbook = openpyxl.load_workbook('AllDataSets.xlsx')
worksheet = workbook['Import Alerts']

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('welcome.html')

@app.route('/search')
def index1():
    firm_name = request.args.get('name')
    resultset = []
    for iterator in range(2, worksheet.max_row + 1):
        try:
            values = worksheet.cell(row=iterator, column=4).value
            link = worksheet.cell(row=iterator, column=4).hyperlink

            client = urlopen(link.display)
            page = client.read()
            client.close()

            page_container = BeautifulSoup(page, 'html.parser')
            container = page_container.findAll('div', {'class': 'div-info'})

            for firms in container:
                content = firms.find_all('div')
                try:
                    if re.search(firm_name, content[0].text, re.IGNORECASE):
                        print("Firm: ", content[0].text, "\n\t found in import alert: ", values, "\n", content[1].text)
                        resultset.append([content[0].text, values, content[1].text])

                except Exception:
                    pass
        except:
            pass

    if len(resultset) > 0:
        return render_template("ResultSet.html",firm=firm_name, resultset=resultset)
    else:
        return render_template("empty.html",firm_name=firm_name)

if __name__ == "__main__":
    app.run(debug=True)
